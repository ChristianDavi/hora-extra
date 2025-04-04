import tkinter as tk
from tkinter import ttk, filedialog
from openpyxl import Workbook, load_workbook
from datetime import datetime

class AppHorasExtras:
    def __init__(self, root):
        self.root = root
        self.root.title('Registro de Horas Extras')
        self.root.geometry('800x600')
        self.root.configure(bg='#C0C0C0')  # Fundo cinza clássico

        self.planilha = 'horas_extras.xlsx'
        self.criar_planilha()

        self.criar_interface()

    def criar_planilha(self):
        try:
            self.wb = load_workbook(self.planilha)
            if 'Registros' not in self.wb.sheetnames:
                self.wb.create_sheet('Registros')
                self.wb.save(self.planilha)
        except FileNotFoundError:
            self.wb = Workbook()
            self.wb.create_sheet('Registros', 0)
            self.wb.save(self.planilha)

    def criar_interface(self):
        frame_principal = tk.Frame(self.root, bg='#C0C0C0', relief=tk.SUNKEN, bd=2)
        frame_principal.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        fonte_padrao = ('MS Sans Serif', 10)

        tk.Label(frame_principal, text="Data (dd/mm/aaaa):", font=fonte_padrao, bg='#C0C0C0').grid(row=1, column=0, sticky='w', pady=5)
        self.data_entry = tk.Entry(frame_principal, font=fonte_padrao, relief=tk.SUNKEN, bd=2)
        self.data_entry.grid(row=1, column=1, pady=5)

        tk.Label(frame_principal, text="Horas Extras (HH:MM):", font=fonte_padrao, bg='#C0C0C0').grid(row=2, column=0, sticky='w', pady=5)
        self.horas_extras_entry = tk.Entry(frame_principal, font=fonte_padrao, relief=tk.SUNKEN, bd=2)
        self.horas_extras_entry.grid(row=2, column=1, pady=5)

        tk.Label(frame_principal, text="Motivo das Horas Extras:", font=fonte_padrao, bg='#C0C0C0').grid(row=3, column=0, sticky='nw', pady=5)
        self.motivo_text = tk.Text(frame_principal, width=50, height=5, font=fonte_padrao, relief=tk.SUNKEN, bd=2)
        self.motivo_text.grid(row=3, column=1, pady=5)

        salvar_btn = tk.Button(frame_principal, text="Salvar", command=self.salvar_registro, font=fonte_padrao, relief=tk.RAISED, bd=3)
        salvar_btn.grid(row=4, column=1, pady=10, sticky='e')

    def salvar_registro(self):
        data = self.data_entry.get()
        horas_extras = self.horas_extras_entry.get()
        motivo = self.motivo_text.get('1.0', tk.END).strip()

        if not data or not horas_extras:
            return

        try:
            planilha = self.wb['Registros']
        except KeyError:
            planilha = self.wb.create_sheet('Registros')

        if planilha.max_row == 1:
            planilha.append(['Data', 'Horas Extras', 'Motivo'])

        nova_linha = [data, horas_extras, motivo]
        planilha.append(nova_linha)
        self.wb.save(self.planilha)

        self.data_entry.delete(0, tk.END)
        self.horas_extras_entry.delete(0, tk.END)
        self.motivo_text.delete('1.0', tk.END)


if __name__ == "__main__":
    root = tk.Tk()
    app = AppHorasExtras(root)
    root.mainloop()
